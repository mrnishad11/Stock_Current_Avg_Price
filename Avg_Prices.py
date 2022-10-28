import openpyxl
import os
from nsepy import get_history
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
date=datetime(2022,10,27)
stock=['AEGISCHEM','BANKBARODA','CANBK','GUJGASLTD','INDIANB','MAHABANK']
ans=[]
for i in stock:
    stock_data=get_history(symbol=i,start=date,end=date)
    stock_data.to_excel(f'{i}.xlsx')
    stockname=i+'.xlsx'
    wb = openpyxl.load_workbook(stockname)
    ws=wb.active    

#VWAP
    for rows in range(1,10):
        copy='J'+str(rows)
        if ws[copy].value==None:
            break
    last_price=ws['J'+str(rows-1)].value
    ans.append(last_price)
    os.remove(i+'.xlsx')
    print(f"Succeessfully Exported...!!!{i}")

Avg=openpyxl.Workbook()
ac=Avg.active
stock.insert(0,'0')
stock.insert(0,'0')
ans.insert(0,'0')
ans.insert(0,'0')
ac['A1']='Stock'
ac['B1']='Avg Price' 
for i in range(2,len(stock)):
    paste1='A'+str(i)
    paste2='B'+str(i)
    ac[paste1]=stock[i]
    ac[paste2]=ans[i]
d=str(date)
Avg.save(f'Avg_price_{d[:10]}.xlsx')    

